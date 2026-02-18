#!/usr/bin/env python3
"""
Corrected ETABS Excel MCP Server - Proper ETABS Database Table Format
Based on actual ETABS table structure and import requirements
"""

import asyncio
import json
import sys
import logging
import math
import os
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

try:
    from mcp.server import Server
    from mcp.server.models import InitializationOptions
    from mcp.server.stdio import stdio_server
    from mcp.types import Tool, TextContent
    
    # Excel generation libraries
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
except ImportError as e:
    logger.error(f"Import error: {e}")
    logger.error("Make sure you have required packages: pip install mcp openpyxl")
    sys.exit(1)

class CorrectedETABSServer:
    def __init__(self):
        try:
            self.server = Server("corrected-etabs-structural-design")
            self.building_data = {}
            self.output_dir = "etabs_exports"
            self.setup_output_directory()
            self.setup_tools()
            logger.info("âœ“ Corrected ETABS MCP Server initialized successfully")
        except Exception as e:
            logger.error(f"âœ— Server initialization failed: {e}")
            raise

    def setup_output_directory(self):
        """Create output directory for exported files"""
        try:
            os.makedirs(self.output_dir, exist_ok=True)
            logger.info(f"âœ“ Output directory ready: {self.output_dir}")
        except Exception as e:
            logger.error(f"âœ— Failed to create output directory: {e}")

    def setup_tools(self):
        """Setup ETABS tools with corrected format"""
        
        @self.server.list_tools()
        async def handle_list_tools() -> List[Tool]:
            return [
                Tool(
                    name="test_connection",
                    description="Test MCP server connection",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "message": {"type": "string", "description": "Test message", "default": "Hello from Corrected ETABS MCP!"}
                        }
                    }
                ),
                
                Tool(
                    name="design_building",
                    description="Design building with correct ETABS format",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "building_name": {"type": "string", "description": "Building name/identifier"},
                            "stories": {"type": "integer", "description": "Number of stories", "minimum": 1, "maximum": 50},
                            "story_height": {"type": "number", "description": "Typical story height (m)", "default": 3.5},
                            "plan_dimensions": {
                                "type": "object",
                                "properties": {
                                    "length": {"type": "number", "description": "Building length (m)"},
                                    "width": {"type": "number", "description": "Building width (m)"}
                                },
                                "required": ["length", "width"]
                            },
                            "bay_spacing": {
                                "type": "object",
                                "properties": {
                                    "x_direction": {"type": "number", "description": "Bay spacing in X direction (m)", "default": 6.0},
                                    "y_direction": {"type": "number", "description": "Bay spacing in Y direction (m)", "default": 6.0}
                                }
                            },
                            "material": {"type": "string", "enum": ["steel", "concrete"], "default": "concrete"}
                        },
                        "required": ["building_name", "stories", "plan_dimensions"]
                    }
                ),
                
                Tool(
                    name="export_etabs_excel_corrected",
                    description="Export to corrected ETABS-compatible Excel format",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "building_name": {"type": "string", "description": "Building name to export"}
                        },
                        "required": ["building_name"]
                    }
                )
            ]

        @self.server.call_tool()
        async def handle_call_tool(name: str, arguments: Dict[str, Any]) -> List[TextContent]:
            try:
                logger.info(f"Tool called: {name} with arguments: {arguments}")
                
                if name == "test_connection":
                    return self._test_connection(arguments)
                elif name == "design_building":
                    return self._design_building(arguments)
                elif name == "export_etabs_excel_corrected":
                    return self._export_etabs_excel_corrected(arguments)
                else:
                    return [TextContent(type="text", text=f"Unknown tool: {name}")]
                    
            except Exception as e:
                error_msg = f"Error in tool '{name}': {str(e)}"
                logger.error(error_msg)
                return [TextContent(type="text", text=error_msg)]

    def _test_connection(self, arguments: Dict[str, Any]) -> List[TextContent]:
        """Test connection tool"""
        message = arguments.get("message", "Hello from Corrected ETABS MCP!")
        response = f"""âœ“ Corrected ETABS MCP Server connected successfully!

Message: {message}

ðŸ”§ FIXES APPLIED:
âœ… Ready for direct import to ETABS 21!

File location: {os.path.abspath(file_path)}"""
        
        return [TextContent(type="text", text=response)]

    def _create_corrected_excel_file(self, building: Dict) -> str:
        """Create Excel file with correct ETABS format"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{building['name']}_ETABS_Corrected_{timestamp}.xlsx"
        file_path = os.path.join(self.output_dir, filename)
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create worksheets with EXACT ETABS table names
        self._create_joint_coordinates_sheet(wb, building)
        self._create_connectivity_frame_sheet(wb, building)
        self._create_material_properties_sheet(wb, building)
        self._create_frame_section_properties_sheet(wb, building)
        
        # Save file
        wb.save(file_path)
        logger.info(f"âœ“ Corrected Excel file created: {file_path}")
        return file_path

    def _create_joint_coordinates_sheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create Joint Coordinates sheet with exact ETABS format"""
        ws = wb.create_sheet("Joint Coordinates")
        
        # EXACT ETABS column headers - based on actual ETABS export format
        headers = [
            "Joint", "CoordSys", "CoordType", "XorR", "Y", "Z", 
            "SpecialJt", "GlobalX", "GlobalY", "GlobalZ"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add joint data in ETABS format
        for joint in building["joints"]:
            row = [
                joint["Joint"],           # Joint name (e.g., "11", "12")
                joint["CoordSys"],        # "GLOBAL"
                joint["CoordType"],       # "Cartesian"
                joint["XorR"],            # X coordinate
                joint["Y"],               # Y coordinate  
                joint["Z"],               # Z coordinate
                joint["SpecialJt"],       # "No"
                joint["GlobalX"],         # Global X (same as XorR)
                joint["GlobalY"],         # Global Y (same as Y)
                joint["GlobalZ"]          # Global Z (same as Z)
            ]
            ws.append(row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_connectivity_frame_sheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create Connectivity - Frame sheet with exact ETABS format"""
        ws = wb.create_sheet("Connectivity - Frame")
        
        # EXACT ETABS column headers
        headers = [
            "Frame", "JointI", "JointJ", "Length", "MatProp", "Section", 
            "Angle", "CardinalPt", "NonPrismatic", "DesignOrient", "GUID"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add frame data
        for frame in building["frames"]:
            row = [
                frame["Frame"],           # Frame name (e.g., "C1", "B1")
                frame["JointI"],          # Start joint
                frame["JointJ"],          # End joint
                frame["Length"],          # Length
                frame["MatProp"],         # Material property
                frame["Section"],         # Section property
                frame["Angle"],           # Angle (0)
                frame.get("CardinalPt", "10-Centroid"),  # Cardinal point
                "No",                     # NonPrismatic
                "Global",                 # DesignOrient
                ""                        # GUID (empty)
            ]
            ws.append(row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_material_properties_sheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create Material Properties 02 - Basic Mechanical Properties sheet"""
        ws = wb.create_sheet("Material Properties 02 - Basic Mechanical Properties")
        
        # EXACT ETABS column headers for material properties
        headers = [
            "Material", "Type", "SymType", "TempDepend", "Color", "Notes", "GUID",
            "E", "U", "A", "G"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add material data
        for mat_name, mat_props in building["materials"].items():
            row = [
                mat_name,                 # Material name
                mat_props["Type"],        # Material type
                "Isotropic",              # Symmetry type
                "No",                     # Temperature dependent
                mat_props.get("Color", "Blue"),  # Color
                "",                       # Notes
                "",                       # GUID
                mat_props["E"],           # Elastic modulus
                mat_props["U"],           # Poisson's ratio
                mat_props["A"],           # Thermal coefficient
                mat_props["G"]            # Shear modulus
            ]
            ws.append(row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_frame_section_properties_sheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create Frame Section Properties 01 - General sheet"""
        ws = wb.create_sheet("Frame Section Properties 01 - General")
        
        # EXACT ETABS column headers for frame sections
        headers = [
            "SectionName", "Material", "Shape", "t3", "t2", "tf", "tw", 
            "Area", "TorsConst", "I33", "I22", "I12", "AS2", "AS3", "S33", "S22", "Z33", "Z22", "R33", "R22"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add section data
        for sect_name, sect_props in building["sections"].items():
            row = [
                sect_name,                        # Section name
                sect_props["Material"],           # Material
                sect_props["Shape"],              # Shape
                sect_props["t3"],                 # Depth (t3)
                sect_props["t2"],                 # Width (t2)
                sect_props.get("tf", 0),          # Flange thickness
                sect_props.get("tw", 0),          # Web thickness
                sect_props["Area"],               # Area
                sect_props["TorsConst"],          # Torsion constant
                sect_props["I33"],                # Moment of inertia 33
                sect_props["I22"],                # Moment of inertia 22
                0,                                # I12 (product of inertia)
                sect_props.get("AS2", 0),         # Shear area 2
                sect_props.get("AS3", 0),         # Shear area 3
                sect_props.get("S33", 0),         # Section modulus 33
                sect_props.get("S22", 0),         # Section modulus 22
                sect_props.get("Z33", 0),         # Plastic section modulus 33
                sect_props.get("Z22", 0),         # Plastic section modulus 22
                sect_props.get("R33", 0),         # Radius of gyration 33
                sect_props.get("R22", 0)          # Radius of gyration 22
            ]
            ws.append(row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _get_etabs_materials(self, material: str) -> Dict:
        """Get materials in ETABS format"""
        if material == "concrete":
            return {
                "CONC": {
                    "Type": "Concrete",
                    "E": 30000000,      # kN/mÂ² (30 GPa)
                    "U": 0.2,           # Poisson's ratio
                    "A": 1e-5,          # Thermal coefficient
                    "G": 12500000,      # Shear modulus
                    "Color": "Gray"
                }
            }
        else:  # steel
            return {
                "STEEL": {
                    "Type": "Steel",
                    "E": 200000000,     # kN/mÂ² (200 GPa)
                    "U": 0.3,           # Poisson's ratio
                    "A": 1.2e-5,        # Thermal coefficient
                    "G": 76923000,      # Shear modulus
                    "Color": "Blue"
                }
            }

    def _get_etabs_sections(self, material: str) -> Dict:
        """Get sections in ETABS format"""
        if material == "concrete":
            return {
                "COL1": {
                    "Material": "CONC",
                    "Shape": "Rectangular",
                    "t3": 0.4,          # Depth (400mm)
                    "t2": 0.4,          # Width (400mm)
                    "Area": 0.16,       # mÂ²
                    "TorsConst": 0.003413,  # mâ´
                    "I33": 0.002133,    # mâ´
                    "I22": 0.002133,    # mâ´
                    "AS2": 0.133,       # Shear area
                    "AS3": 0.133,       # Shear area
                    "S33": 0.010667,    # Section modulus
                    "S22": 0.010667,    # Section modulus
                    "Z33": 0.016,       # Plastic section modulus
                    "Z22": 0.016,       # Plastic section modulus
                    "R33": 0.115,       # Radius of gyration
                    "R22": 0.115        # Radius of gyration
                },
                "BEAM1": {
                    "Material": "CONC",
                    "Shape": "Rectangular",
                    "t3": 0.6,          # Depth (600mm)
                    "t2": 0.3,          # Width (300mm)
                    "Area": 0.18,       # mÂ²
                    "TorsConst": 0.002808,  # mâ´
                    "I33": 0.0054,      # mâ´
                    "I22": 0.00135,     # mâ´
                    "AS2": 0.15,        # Shear area
                    "AS3": 0.15,        # Shear area
                    "S33": 0.018,       # Section modulus
                    "S22": 0.009,       # Section modulus
                    "Z33": 0.027,       # Plastic section modulus
                    "Z22": 0.0135,      # Plastic section modulus
                    "R33": 0.173,       # Radius of gyration
                    "R22": 0.087        # Radius of gyration
                }
            }
        else:  # steel
            return {
                "COL1": {
                    "Material": "STEEL",
                    "Shape": "I/Wide Flange",
                    "t3": 0.24,         # Depth (HEB240)
                    "t2": 0.24,         # Width
                    "tf": 0.017,        # Flange thickness
                    "tw": 0.01,         # Web thickness
                    "Area": 0.0106,     # mÂ²
                    "TorsConst": 0.0000618,  # mâ´
                    "I33": 0.00011259,  # mâ´
                    "I22": 0.00003923,  # mâ´
                    "AS2": 0.0024,      # Shear area
                    "AS3": 0.0024,      # Shear area
                    "S33": 0.000938,    # Section modulus
                    "S22": 0.000327,    # Section modulus
                    "Z33": 0.001056,    # Plastic section modulus
                    "Z22": 0.000491,    # Plastic section modulus
                    "R33": 0.103,       # Radius of gyration
                    "R22": 0.061        # Radius of gyration
                },
                "BEAM1": {
                    "Material": "STEEL",
                    "Shape": "I/Wide Flange",
                    "t3": 0.36,         # Depth (IPE360)
                    "t2": 0.17,         # Width
                    "tf": 0.0127,       # Flange thickness
                    "tw": 0.008,        # Web thickness
                    "Area": 0.0073,     # mÂ²
                    "TorsConst": 0.0000377,  # mâ´
                    "I33": 0.0001627,   # mâ´
                    "I22": 0.000001043, # mâ´
                    "AS2": 0.002016,    # Shear area
                    "AS3": 0.002016,    # Shear area
                    "S33": 0.000904,    # Section modulus
                    "S22": 0.000123,    # Section modulus
                    "Z33": 0.001019,    # Plastic section modulus
                    "Z22": 0.000189,    # Plastic section modulus
                    "R33": 0.149,       # Radius of gyration
                    "R22": 0.038        # Radius of gyration
                }
            }

async def main():
    """Main server entry point"""
    try:
        logger.info("Starting Corrected ETABS MCP server...")
        server_instance = CorrectedETABSServer()
        
        async with stdio_server() as (read_stream, write_stream):
            logger.info("Corrected ETABS server running and waiting for connections...")
            
            try:
                await server_instance.server.run(
                    read_stream,
                    write_stream,
                    InitializationOptions(
                        server_name="corrected-etabs-structural-design",
                        server_version="5.0.0",
                        capabilities={}
                    ),
                )
            except asyncio.CancelledError:
                logger.info("Server cancelled")
                raise
            except Exception as e:
                logger.error(f"Error during server run: {e}")
                raise
                
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Server shutdown")
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        sys.exit(1) Correct ETABS database table names
âœ… Proper column headers and formats
âœ… ETABS-compatible data structure
âœ… Based on actual ETABS Excel export format

ðŸ“ Output Directory: {os.path.abspath(self.output_dir)}
ðŸŽ¯ Ready for accurate ETABS import!"""
        return [TextContent(type="text", text=response)]

    def _design_building(self, arguments: Dict[str, Any]) -> List[TextContent]:
        """Design building with proper ETABS structure"""
        building_name = arguments["building_name"]
        stories = arguments["stories"]
        story_height = arguments.get("story_height", 3.5)
        plan_dims = arguments["plan_dimensions"]
        bay_spacing = arguments.get("bay_spacing", {"x_direction": 6.0, "y_direction": 6.0})
        material = arguments.get("material", "concrete")
        
        # Calculate building geometry
        length = plan_dims["length"]
        width = plan_dims["width"]
        total_height = stories * story_height
        
        # Calculate grid lines
        x_bays = math.ceil(length / bay_spacing["x_direction"])
        y_bays = math.ceil(width / bay_spacing["y_direction"])
        
        actual_x_spacing = length / x_bays
        actual_y_spacing = width / y_bays
        
        # Generate joints with ETABS naming convention
        joints = []
        joint_id = 1
        for story in range(stories + 1):  # Include base level
            z = story * story_height
            for i in range(x_bays + 1):
                for j in range(y_bays + 1):
                    # ETABS style joint naming
                    joint_name = f"{i+1}{j+1}"  # Grid-based naming like "11", "12", etc.
                    
                    joints.append({
                        "Joint": joint_name,
                        "CoordSys": "GLOBAL",
                        "CoordType": "Cartesian", 
                        "XorR": round(i * actual_x_spacing, 3),
                        "Y": round(j * actual_y_spacing, 3),
                        "Z": round(z, 3),
                        "SpecialJt": "No",
                        "GlobalX": round(i * actual_x_spacing, 3),
                        "GlobalY": round(j * actual_y_spacing, 3),
                        "GlobalZ": round(z, 3),
                        "story": story,
                        "grid_i": i,
                        "grid_j": j
                    })
                    joint_id += 1
        
        # Generate frames with ETABS naming
        frames = []
        frame_id = 1
        points_per_story = (x_bays + 1) * (y_bays + 1)
        
        # Columns
        for story in range(stories):
            for i in range(x_bays + 1):
                for j in range(y_bays + 1):
                    # Find joints for this column
                    joint_bottom = f"{i+1}{j+1}"
                    joint_top = f"{i+1}{j+1}"  # Same grid position, different story
                    
                    # ETABS style frame naming
                    frame_name = f"C{frame_id}"
                    
                    frames.append({
                        "Frame": frame_name,
                        "JointI": joint_bottom,
                        "JointJ": joint_top,  
                        "Length": story_height,
                        "MatProp": "CONC" if material == "concrete" else "STEEL",
                        "Section": "COL1",
                        "Angle": 0,
                        "CardinalPt": "10-Centroid",
                        "type": "column"
                    })
                    frame_id += 1
        
        # Beams - X direction
        for story in range(1, stories + 1):
            for j in range(y_bays + 1):
                for i in range(x_bays):
                    joint_i = f"{i+1}{j+1}"
                    joint_j = f"{i+2}{j+1}"
                    frame_name = f"B{frame_id}"
                    
                    frames.append({
                        "Frame": frame_name,
                        "JointI": joint_i,
                        "JointJ": joint_j,
                        "Length": actual_x_spacing,
                        "MatProp": "CONC" if material == "concrete" else "STEEL",
                        "Section": "BEAM1",
                        "Angle": 0,
                        "CardinalPt": "10-Centroid",
                        "type": "beam"
                    })
                    frame_id += 1
        
        # Beams - Y direction  
        for story in range(1, stories + 1):
            for i in range(x_bays + 1):
                for j in range(y_bays):
                    joint_i = f"{i+1}{j+1}"
                    joint_j = f"{i+1}{j+2}"
                    frame_name = f"B{frame_id}"
                    
                    frames.append({
                        "Frame": frame_name,
                        "JointI": joint_i,
                        "JointJ": joint_j,
                        "Length": actual_y_spacing,
                        "MatProp": "CONC" if material == "concrete" else "STEEL",
                        "Section": "BEAM1",
                        "Angle": 0,
                        "CardinalPt": "10-Centroid",
                        "type": "beam"
                    })
                    frame_id += 1
        
        # Generate materials in ETABS format
        materials = self._get_etabs_materials(material)
        sections = self._get_etabs_sections(material)
        
        # Store building data
        building_data = {
            "name": building_name,
            "geometry": {
                "stories": stories,
                "story_height": story_height,
                "total_height": total_height,
                "plan_dimensions": {"length": length, "width": width},
                "bay_spacing": {"x": actual_x_spacing, "y": actual_y_spacing}
            },
            "material": material,
            "joints": joints,
            "frames": frames,
            "materials": materials,
            "sections": sections,
            "design_summary": {
                "total_joints": len(joints),
                "total_frames": len(frames),
                "columns": len([f for f in frames if f["type"] == "column"]),
                "beams": len([f for f in frames if f["type"] == "beam"])
            }
        }
        
        self.building_data[building_name] = building_data
        
        result = {
            "building_summary": {
                "name": building_name,
                "dimensions": f"{length}m x {width}m x {total_height}m",
                "stories": stories,
                "material": material
            },
            "model_statistics": building_data["design_summary"],
            "etabs_format": "Corrected to match actual ETABS database tables",
            "next_steps": [
                f"Use 'export_etabs_excel_corrected' to generate proper Excel file",
                f"File will use correct ETABS table names and column headers"
            ]
        }
        
        response = f"Building Design Complete - ETABS Format Corrected:\n\n{json.dumps(result, indent=2)}"
        return [TextContent(type="text", text=response)]

    def _export_etabs_excel_corrected(self, arguments: Dict[str, Any]) -> List[TextContent]:
        """Export with corrected ETABS format"""
        building_name = arguments["building_name"]
        
        if building_name not in self.building_data:
            return [TextContent(type="text", text=f"Building '{building_name}' not found. Please design it first.")]
        
        building = self.building_data[building_name]
        
        # Create Excel file with correct ETABS format
        file_path = self._create_corrected_excel_file(building)
        
        response = f"""âœ… CORRECTED ETABS Excel File Generated!

ðŸ“ File Details:
  â€¢ File Name: {os.path.basename(file_path)}
  â€¢ Full Path: {os.path.abspath(file_path)}
  â€¢ File Size: {os.path.getsize(file_path) / 1024:.1f} KB
  â€¢ Format: Corrected ETABS database table structure

ðŸ“Š Model Statistics:
  â€¢ Building: {building['name']}
  â€¢ Joints: {building['design_summary']['total_joints']}
  â€¢ Frames: {building['design_summary']['total_frames']}

ðŸ”§ CORRECTIONS APPLIED:
  âœ… Proper table names (e.g., "Joint Coordinates", "Connectivity - Frame")
  âœ… Correct column headers matching ETABS exactly
  âœ… ETABS-style joint naming (grid-based)
  âœ… Proper data types and formatting
  âœ… Compatible with ETABS 21 import process

ðŸš€ ETABS 21 Import Steps:
1. Open ETABS 21
2. File > Import > From Excel Database Tables
3. Select: {os.path.basename(file_path)}
4. Map tables automatically (names should match)
5. Import and verify

ðŸ“‹ Excel Worksheets:
  â€¢ Joint Coordinates - Joint locations and properties
  â€¢ Connectivity - Frame - Frame connectivity and properties
  â€¢ Material Properties 02 - Basic Mechanical Properties
  â€¢ Frame Section Properties 01 - General

âœ…