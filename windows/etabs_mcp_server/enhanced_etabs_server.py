#!/usr/bin/env python3
"""
Enhanced ETABS 21 MCP Server with Excel File Generation
Generates actual .xlsx files for direct ETABS import
"""

import asyncio
import json
import sys
import logging
import math
import os
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

try:
    from mcp.server import Server
    from mcp.server.models import InitializationOptions
    from mcp.server.stdio import stdio_server
    from mcp.types import Tool, TextContent
    
    # Excel generation libraries
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
except ImportError as e:
    logger.error(f"Import error: {e}")
    logger.error("Make sure you have required packages: pip install mcp openpyxl")
    sys.exit(1)

class EnhancedETABSServer:
    def __init__(self):
        try:
            self.server = Server("enhanced-etabs-structural-design")
            self.building_data = {}  # Store building information
            self.output_dir = "etabs_exports"  # Output directory for files
            self.setup_output_directory()
            self.setup_tools()
            logger.info("✓ Enhanced ETABS MCP Server initialized successfully")
        except Exception as e:
            logger.error(f"✗ Server initialization failed: {e}")
            raise

    def setup_output_directory(self):
        """Create output directory for exported files"""
        try:
            os.makedirs(self.output_dir, exist_ok=True)
            logger.info(f"✓ Output directory ready: {self.output_dir}")
        except Exception as e:
            logger.error(f"✗ Failed to create output directory: {e}")

    def setup_tools(self):
        """Setup all ETABS 21 compatible tools with Excel file generation"""
        
        @self.server.list_tools()
        async def handle_list_tools() -> List[Tool]:
            return [
                Tool(
                    name="test_connection",
                    description="Test MCP server connection",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "message": {"type": "string", "description": "Test message", "default": "Hello from Enhanced ETABS MCP!"}
                        }
                    }
                ),
                
                Tool(
                    name="design_building",
                    description="Design a complete building structure for ETABS 21",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "building_name": {"type": "string", "description": "Building name/identifier"},
                            "stories": {"type": "integer", "description": "Number of stories", "minimum": 1, "maximum": 50},
                            "story_height": {"type": "number", "description": "Typical story height (m)", "default": 3.5},
                            "plan_dimensions": {
                                "type": "object",
                                "properties": {
                                    "length": {"type": "number", "description": "Building length (m)"},
                                    "width": {"type": "number", "description": "Building width (m)"}
                                },
                                "required": ["length", "width"]
                            },
                            "bay_spacing": {
                                "type": "object",
                                "properties": {
                                    "x_direction": {"type": "number", "description": "Bay spacing in X direction (m)", "default": 6.0},
                                    "y_direction": {"type": "number", "description": "Bay spacing in Y direction (m)", "default": 6.0}
                                }
                            },
                            "loads": {
                                "type": "object",
                                "properties": {
                                    "dead_load": {"type": "number", "description": "Dead load (kN/m2)", "default": 3.0},
                                    "live_load": {"type": "number", "description": "Live load (kN/m2)", "default": 2.5},
                                    "wind_load": {"type": "number", "description": "Wind load (kN/m2)", "default": 1.0}
                                }
                            },
                            "structural_system": {"type": "string", "enum": ["moment_frame", "braced_frame", "shear_wall"], "default": "moment_frame"},
                            "material": {"type": "string", "enum": ["steel", "concrete", "composite"], "default": "concrete"}
                        },
                        "required": ["building_name", "stories", "plan_dimensions"]
                    }
                ),
                
                Tool(
                    name="export_etabs21_excel",
                    description="Export to ETABS 21 compatible Excel file (.xlsx)",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "building_name": {"type": "string", "description": "Building name to export"},
                            "save_file": {"type": "boolean", "description": "Save actual .xlsx file", "default": True},
                            "file_path": {"type": "string", "description": "Custom file path (optional)"},
                            "table_types": {
                                "type": "array",
                                "items": {"type": "string", "enum": ["joints", "frames", "materials", "sections", "loads", "all"]},
                                "default": ["all"]
                            }
                        },
                        "required": ["building_name"]
                    }
                ),
                
                Tool(
                    name="list_exported_files",
                    description="List all exported ETABS files",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "file_type": {"type": "string", "enum": ["xlsx", "all"], "default": "all"}
                        }
                    }
                ),
                
                Tool(
                    name="generate_import_instructions",
                    description="Generate detailed ETABS import instructions for exported files",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "file_name": {"type": "string", "description": "Exported file name"}
                        },
                        "required": ["file_name"]
                    }
                )
            ]

        @self.server.call_tool()
        async def handle_call_tool(name: str, arguments: Dict[str, Any]) -> List[TextContent]:
            try:
                logger.info(f"Tool called: {name} with arguments: {arguments}")
                
                if name == "test_connection":
                    return self._test_connection(arguments)
                elif name == "design_building":
                    return self._design_building(arguments)
                elif name == "export_etabs21_excel":
                    return self._export_etabs21_excel(arguments)
                elif name == "list_exported_files":
                    return self._list_exported_files(arguments)
                elif name == "generate_import_instructions":
                    return self._generate_import_instructions(arguments)
                else:
                    return [TextContent(type="text", text=f"Unknown tool: {name}")]
                    
            except Exception as e:
                error_msg = f"Error in tool '{name}': {str(e)}"
                logger.error(error_msg)
                return [TextContent(type="text", text=error_msg)]

    def _test_connection(self, arguments: Dict[str, Any]) -> List[TextContent]:
        """Test connection tool"""
        message = arguments.get("message", "Hello from Enhanced ETABS MCP!")
        
        # Check output directory status
        files_count = len([f for f in os.listdir(self.output_dir) if f.endswith('.xlsx')]) if os.path.exists(self.output_dir) else 0
        
        response = f"""✓ Enhanced ETABS MCP Server connected successfully!

Message: {message}

Server Status: Running
Tools Available: 5
Output Directory: {os.path.abspath(self.output_dir)}
Excel Files Generated: {files_count}

🏗️ ENHANCED FEATURES:
✅ Building Design - Complete structural models
✅ Excel File Generation - Actual .xlsx files for ETABS import
✅ File Management - List and organize exported files
✅ Import Instructions - Step-by-step ETABS import guides
✅ Automatic File Save - Files saved to: {self.output_dir}/

📋 READY FOR STRUCTURAL DESIGN!"""
        return [TextContent(type="text", text=response)]

    def _design_building(self, arguments: Dict[str, Any]) -> List[TextContent]:
        """Design a complete building structure"""
        building_name = arguments["building_name"]
        stories = arguments["stories"]
        story_height = arguments.get("story_height", 3.5)
        plan_dims = arguments["plan_dimensions"]
        bay_spacing = arguments.get("bay_spacing", {"x_direction": 6.0, "y_direction": 6.0})
        loads = arguments.get("loads", {"dead_load": 3.0, "live_load": 2.5, "wind_load": 1.0})
        structural_system = arguments.get("structural_system", "moment_frame")
        material = arguments.get("material", "concrete")
        
        # Calculate building geometry
        length = plan_dims["length"]
        width = plan_dims["width"]
        total_height = stories * story_height
        
        # Calculate grid lines
        x_bays = math.ceil(length / bay_spacing["x_direction"])
        y_bays = math.ceil(width / bay_spacing["y_direction"])
        
        actual_x_spacing = length / x_bays
        actual_y_spacing = width / y_bays
        
        # Generate ETABS-compatible joint data
        joints = []
        joint_id = 1
        for story in range(stories + 1):  # Include base level
            z = story * story_height
            for i in range(x_bays + 1):
                for j in range(y_bays + 1):
                    x = i * actual_x_spacing
                    y = j * actual_y_spacing
                    joints.append({
                        "name": f"J{joint_id}",
                        "x": round(x, 3),
                        "y": round(y, 3),
                        "z": round(z, 3),
                        "grid_id": f"G{i+1}-{j+1}",
                        "story": story
                    })
                    joint_id += 1
        
        # Generate frame elements
        frames = []
        frame_id = 1
        points_per_story = (x_bays + 1) * (y_bays + 1)
        
        # Columns
        for story in range(stories):
            for i in range(points_per_story):
                j1_idx = story * points_per_story + i
                j2_idx = (story + 1) * points_per_story + i
                frames.append({
                    "name": f"C{frame_id}",
                    "type": "column",
                    "joint1": joints[j1_idx]["name"],
                    "joint2": joints[j2_idx]["name"],
                    "section": self._get_member_section(material, "column", loads),
                    "material": material.upper(),
                    "story": story + 1
                })
                frame_id += 1
        
        # Beams (simplified - main grid only)
        for story in range(1, stories + 1):
            story_base = story * points_per_story
            
            # X-direction beams
            for j in range(y_bays + 1):
                for i in range(x_bays):
                    j1_idx = story_base + j * (x_bays + 1) + i
                    j2_idx = story_base + j * (x_bays + 1) + i + 1
                    frames.append({
                        "name": f"BX{frame_id}",
                        "type": "beam",
                        "joint1": joints[j1_idx]["name"],
                        "joint2": joints[j2_idx]["name"],
                        "section": self._get_member_section(material, "beam", loads),
                        "material": material.upper(),
                        "story": story
                    })
                    frame_id += 1
            
            # Y-direction beams
            for i in range(x_bays + 1):
                for j in range(y_bays):
                    j1_idx = story_base + j * (x_bays + 1) + i
                    j2_idx = story_base + (j + 1) * (x_bays + 1) + i
                    frames.append({
                        "name": f"BY{frame_id}",
                        "type": "beam", 
                        "joint1": joints[j1_idx]["name"],
                        "joint2": joints[j2_idx]["name"],
                        "section": self._get_member_section(material, "beam", loads),
                        "material": material.upper(),
                        "story": story
                    })
                    frame_id += 1
        
        # Generate materials and sections
        materials = self._get_material_properties(material)
        sections = self._get_section_properties(material, loads)
        load_patterns = self._get_load_patterns(loads)
        
        # Store building data
        building_data = {
            "name": building_name,
            "geometry": {
                "stories": stories,
                "story_height": story_height,
                "total_height": total_height,
                "plan_dimensions": {"length": length, "width": width},
                "bay_spacing": {"x": actual_x_spacing, "y": actual_y_spacing},
                "x_bays": x_bays,
                "y_bays": y_bays
            },
            "loads": loads,
            "structural_system": structural_system,
            "material": material,
            "joints": joints,  
            "frames": frames,
            "materials": materials,
            "sections": sections,
            "load_patterns": load_patterns,
            "design_summary": {
                "total_joints": len(joints),
                "total_frames": len(frames),
                "columns": len([f for f in frames if f["type"] == "column"]),
                "beams": len([f for f in frames if f["type"] == "beam"])
            }
        }
        
        self.building_data[building_name] = building_data
        
        result = {
            "building_summary": {
                "name": building_name,
                "dimensions": f"{length}m x {width}m x {total_height}m",
                "stories": stories,
                "structural_system": structural_system,
                "material": material
            },
            "model_statistics": building_data["design_summary"],
            "grid_layout": {
                "x_bays": x_bays,
                "y_bays": y_bays,
                "actual_spacing": {"x": actual_x_spacing, "y": actual_y_spacing}
            },
            "etabs21_export_ready": True,
            "next_steps": [
                f"Use 'export_etabs21_excel' to generate .xlsx file for ETABS import",
                f"Files will be saved to: {self.output_dir}/",
                f"Use 'generate_import_instructions' for detailed ETABS import steps"
            ]
        }
        
        response = f"Building Design Complete - Ready for ETABS 21:\n\n{json.dumps(result, indent=2)}"
        return [TextContent(type="text", text=response)]

    def _export_etabs21_excel(self, arguments: Dict[str, Any]) -> List[TextContent]:
        """Export to actual ETABS 21 compatible Excel file"""
        building_name = arguments["building_name"]
        save_file = arguments.get("save_file", True)
        custom_path = arguments.get("file_path")
        table_types = arguments.get("table_types", ["all"])
        
        if building_name not in self.building_data:
            return [TextContent(type="text", text=f"Building '{building_name}' not found. Please design it first using 'design_building'.")]
        
        building = self.building_data[building_name]
        
        if save_file:
            # Generate actual Excel file
            file_path = self._create_excel_file(building, custom_path)
            
            response = f"""✅ ETABS 21 Excel File Generated Successfully!

📁 File Details:
  • File Name: {os.path.basename(file_path)}
  • Full Path: {os.path.abspath(file_path)}
  • File Size: {os.path.getsize(file_path) / 1024:.1f} KB
  • Created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

📊 Model Statistics:
  • Building: {building['name']}
  • Joints: {building['design_summary']['total_joints']}
  • Frames: {building['design_summary']['total_frames']}
  • Columns: {building['design_summary']['columns']}
  • Beams: {building['design_summary']['beams']}

🚀 ETABS 21 Import Instructions:
1. Open ETABS 21
2. File > Import > From Excel Database Tables
3. Browse to: {os.path.abspath(file_path)}
4. Select appropriate table mappings
5. Import and verify model

📋 Excel Worksheets Created:
  • Joint_Coordinates - All joint locations
  • Frame_Connectivity - All frame elements  
  • Material_Properties - Material definitions
  • Frame_Sections - Section properties
  • Load_Patterns - Load case definitions

✅ Ready for direct import to ETABS 21!

Use 'generate_import_instructions' for detailed step-by-step guide."""
            
        else:
            # Return text format only
            response = self._generate_excel_text_format(building)
        
        return [TextContent(type="text", text=response)]

    def _create_excel_file(self, building: Dict, custom_path: Optional[str] = None) -> str:
        """Create actual Excel file for ETABS import"""
        
        # Determine file path
        if custom_path:
            file_path = custom_path
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{building['name']}_ETABS21_{timestamp}.xlsx"
            file_path = os.path.join(self.output_dir, filename)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create worksheets
        self._create_joints_worksheet(wb, building)
        self._create_frames_worksheet(wb, building)
        self._create_materials_worksheet(wb, building)
        self._create_sections_worksheet(wb, building)
        self._create_loads_worksheet(wb, building)
        
        # Save file
        wb.save(file_path)
        
        logger.info(f"✓ Excel file created: {file_path}")
        return file_path

    def _create_joints_worksheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create joints worksheet"""
        ws = wb.create_sheet("Joint_Coordinates")
        
        # Headers
        headers = ["Joint", "CoordSys", "CoordType", "XorR", "Y", "Z", "SpecialJt", "GlobalX", "GlobalY", "GlobalZ"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add joint data
        for joint in building["joints"]:
            row = [
                joint["name"], "GLOBAL", "Cartesian",
                joint["x"], joint["y"], joint["z"], "No",
                joint["x"], joint["y"], joint["z"]
            ]
            ws.append(row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_frames_worksheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create frames worksheet"""
        ws = wb.create_sheet("Frame_Connectivity")
        
        # Headers  
        headers = ["Frame", "JointI", "JointJ", "Length", "MatProp", "Section", "Angle"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add frame data
        for frame in building["frames"]:
            # Calculate length
            j1 = next(j for j in building["joints"] if j["name"] == frame["joint1"])
            j2 = next(j for j in building["joints"] if j["name"] == frame["joint2"])
            length = round(math.sqrt((j2["x"]-j1["x"])**2 + (j2["y"]-j1["y"])**2 + (j2["z"]-j1["z"])**2), 3)
            
            row = [
                frame["name"], frame["joint1"], frame["joint2"],
                length, frame["material"], frame["section"], 0
            ]
            ws.append(row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_materials_worksheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create materials worksheet"""
        ws = wb.create_sheet("Material_Properties")
        
        # Headers
        headers = ["Material", "Type", "SymType", "TempDepend", "Color", "Notes", "GUID", "E", "U", "A", "G"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add material data
        for mat_name, mat_props in building["materials"].items():
            row = [
                mat_name, mat_props["type"], "Isotropic", "No", "Blue", "", "",
                mat_props["E"], mat_props["poisson"], mat_props["thermal"], mat_props["G"]
            ]
            ws.append(row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_sections_worksheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create sections worksheet"""
        ws = wb.create_sheet("Frame_Sections")
        
        # Headers
        headers = ["SectionName", "Material", "Shape", "t3", "t2", "tf", "tw", "Area", "TorsConst", "I33", "I22"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add section data
        for sect_name, sect_props in building["sections"].items():
            row = [
                sect_name, sect_props["material"], sect_props["shape"],
                sect_props["depth"], sect_props["width"], 
                sect_props.get("flange_thickness", 0), sect_props.get("web_thickness", 0),
                sect_props["area"], sect_props["J"], sect_props["I33"], sect_props["I22"]
            ]
            ws.append(row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_loads_worksheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create load patterns worksheet"""
        ws = wb.create_sheet("Load_Patterns")
        
        # Headers
        headers = ["LoadPat", "DesignType", "Type", "SelfWtMult", "AutoLoad", "Notes", "GUID"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add load data
        for load_pat in building["load_patterns"]:
            row = [
                load_pat["name"], load_pat["design_type"], load_pat["type"],
                load_pat["self_weight"], load_pat["auto_load"], load_pat["notes"], ""
            ]
            ws.append(row)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _list_exported_files(self, arguments: Dict[str, Any]) -> List[TextContent]:
        """List all exported files"""
        file_type = arguments.get("file_type", "all")
        
        if not os.path.exists(self.output_dir):
            return [TextContent(type="text", text=f"Output directory does not exist: {self.output_dir}")]
        
        files = []
        for filename in os.listdir(self.output_dir):
            file_path = os.path.join(self.output_dir, filename)
            if os.path.isfile(file_path):
                if file_type == "all" or filename.endswith(f".{file_type}"):
                    file_size = os.path.getsize(file_path)
                    file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                    files.append({
                        "name": filename,
                        "size_kb": round(file_size / 1024, 1),
                        "modified": file_time.strftime("%Y-%m-%d %H:%M:%S"),
                        "path": os.path.abspath(file_path)
                    })
        
        if not files:
            response = f"No files found in {self.output_dir}"
        else:
            response = f"📁 Exported Files ({len(files)} found):\n"
            response += f"Directory: {os.path.abspath(self.output_dir)}\n\n"
            
            for file_info in sorted(files, key=lambda x: x["modified"], reverse=True):
                response += f"📄 {file_info['name']}\n"
                response += f"   Size: {file_info['size_kb']} KB\n"
                response += f"   Modified: {file_info['modified']}\n"
                response += f"   Path: {file_info['path']}\n\n"
        
        return [TextContent(type="text", text=response)]

    def _generate_import_instructions(self, arguments: Dict[str, Any]) -> List[TextContent]:
        """Generate detailed import instructions"""
        file_name = arguments["file_name"]
        file_path = os.path.join(self.output_dir, file_name)
        
        if not os.path.exists(file_path):
            return [TextContent(type="text", text=f"File not found: {file_name}")]
        
        instructions = f"""📋 ETABS 21 Import Instructions for {file_name}
{'='*80}

📁 FILE LOCATION:
{os.path.abspath(file_path)}

🚀 STEP-BY-STEP IMPORT PROCESS:

1️⃣ PREPARE ETABS:
   • Open ETABS 21
   • Create new blank model (or open existing)
   • Set units to kN, m, C (Analyze > Set Units)

2️⃣ START IMPORT:
   • Go to: File > Import > From Excel Database Tables
   • Browse and select: {file_name}
   • Click "Open"

3️⃣ TABLE MAPPING:
   ETABS will show "Choose Tables" dialog. Map as follows:
   
   Excel Sheet                    →  ETABS Table
   ----------------------------------------
   Joint_Coordinates             →  Joint Coordinates
   Frame_Connectivity            →  Connectivity - Frame
   Material_Properties           →  Material Properties 02 - Basic Mechanical Properties
   Frame_Sections               →  Frame Section Properties 01 - General  
   Load_Patterns                →  Load Pattern Definitions

4️⃣ IMPORT EXECUTION:
   • Check all required tables
   • Click "OK" to import
   • Watch for any error messages
   • ETABS will process and import data

5️⃣ VERIFICATION:
   • Press F4 to view 3D model
   • Check structure appears correctly
   • Verify joint coordinates (Select > By Joint > All)
   • Check frame connectivity (Select > By Frame > All)
   • Review materials (Define > Material Properties)
   • Confirm sections (Define > Section Properties > Frame Sections)

6️⃣ COMPLETE MODEL SETUP:
   • Add support restraints to base joints:
     - Assign > Joint > Restraints
     - Select base joints
     - Apply Fixed or Pinned supports
   
   • Define load combinations:
     - Define > Load Combinations
     - Add: 1.2DL + 1.6LL (Ultimate)
     - Add: 1.0DL + 1.0LL (Service)
   
   • Apply loads to frames:
     - Assign > Frame Loads > Distributed
     - Apply dead and live loads as needed

7️⃣ RUN ANALYSIS:
   • Analyze > Run Analysis
   • Check for warnings/errors
   • Review results

✅ SUCCESS INDICATORS:
   • 3D model displays properly
   • All joints and frames imported
   • Materials and sections assigned
   • No import error messages

⚠️ TROUBLESHOOTING:
   • If import fails: Check Excel file is not open in another program
   • Missing data: Verify all required columns have data
   • Wrong units: Set ETABS units before import
   • Connection issues: Check joint names match in frame connectivity

📞 SUPPORT:
   • Use ETABS Help > Documentation for detailed import procedures
   • Check modelcontextprotocol.io for MCP server updates

File imported: {os.path.basename(file_path)}
File size: {os.path.getsize(file_path) / 1024:.1f} KB
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Ready for structural analysis in ETABS 21! 🏗️"""
        
        return [TextContent(type="text", text=instructions)]

    def _generate_excel_text_format(self, building: Dict) -> str:
        """Generate text format for preview (when save_file=False)"""
        return f"""ETABS 21 Excel Export Preview - {building['name']}
{'='*60}

This is a preview. Use save_file=true to generate actual .xlsx file.

JOINT COORDINATES ({len(building['joints'])} joints):
Joint\tCoordSys\tCoordType\tXorR\tY\tZ
{building['joints'][0]['name']}\tGLOBAL\tCartesian\t{building['joints'][0]['x']}\t{building['joints'][0]['y']}\t{building['joints'][0]['z']}
... (showing first joint only)

FRAME CONNECTIVITY ({len(building['frames'])} frames):
Frame\tJointI\tJointJ\tLength\tMatProp\tSection
{building['frames'][0]['name']}\t{building['frames'][0]['joint1']}\t{building['frames'][0]['joint2']}\t...\t{building['frames'][0]['material']}\t{building['frames'][0]['section']}
... (showing first frame only)

Use export_etabs21_excel with save_file=true to generate complete .xlsx file."""

    # Helper methods (same as before)
    def _get_member_section(self, material: str, member_type: str, loads: Dict) -> str:
        """Get appropriate member section based on material and loads"""
        load_intensity = loads["dead_load"] + loads["live_load"]
        
        if material == "concrete":
            if member_type == "column":
                if load_intensity < 4:
                    return "COL_300x300"
                elif load_intensity < 6:
                    return "COL_400x400"
                else:
                    return "COL_500x500"
            else:  # beam
                if load_intensity < 4:
                    return "BEAM_300x500"
                elif load_intensity < 6:
                    return "BEAM_300x600"
                else:
                    return "BEAM_400x700"
        else:  # steel
            if member_type == "column":
                if load_intensity < 4:
                    return "HEB200"
                elif load_intensity < 6:
                    return "HEB240"
                else:
                    return "HEB300"
            else:  # beam
                if load_intensity < 4:
                    return "IPE300"
                elif load_intensity < 6:
                    return "IPE360"
                else:
                    return "IPE450"

    def _get_material_properties(self, material: str) -> Dict:
        """Get material properties dictionary"""
        if material == "concrete":
            return {
                "CONCRETE": {
                    "type": "Concrete",
                    "E": 30000000,  # kN/m2 (30 GPa)
                    "poisson": 0.2,
                    "thermal": 1e-5,
                    "G": 12500000,
                    "density": 2400  # kg/m3
                }
            }
        elif material == "steel":
            return {
                "STEEL": {
                    "type": "Steel",  
                    "E": 200000000,  # kN/m2 (200 GPa)
                    "poisson": 0.3,
                    "thermal": 1.2e-5,
                    "G": 76923000,
                    "density": 7850  # kg/m3
                }
            }
        else:  # composite
            return {
                "STEEL": {
                    "type": "Steel",
                    "E": 200000000,
                    "poisson": 0.3,
                    "thermal": 1.2e-5,
                    "G": 76923000,
                    "density": 7850
                },
                "CONCRETE": {
                    "type": "Concrete",
                    "E": 30000000,
                    "poisson": 0.2,
                    "thermal": 1e-5,
                    "G": 12500000,
                    "density": 2400
                }
            }

    def _get_section_properties(self, material: str, loads: Dict) -> Dict:
        """Get section properties dictionary"""
        sections = {}
        
        if material == "concrete":
            # Concrete sections
            sections.update({
                "COL_300x300": {
                    "material": "CONCRETE",
                    "shape": "Rectangle",
                    "width": 0.3,
                    "depth": 0.3,
                    "area": 0.09,
                    "I33": 0.000675,  # m4
                    "I22": 0.000675,
                    "J": 0.00108
                },
                "COL_400x400": {
                    "material": "CONCRETE", 
                    "shape": "Rectangle",
                    "width": 0.4,
                    "depth": 0.4,
                    "area": 0.16,
                    "I33": 0.002133,
                    "I22": 0.002133,
                    "J": 0.003413
                },
                "COL_500x500": {
                    "material": "CONCRETE",
                    "shape": "Rectangle", 
                    "width": 0.5,
                    "depth": 0.5,
                    "area": 0.25,
                    "I33": 0.005208,
                    "I22": 0.005208,
                    "J": 0.008333
                },
                "BEAM_300x500": {
                    "material": "CONCRETE",
                    "shape": "Rectangle",
                    "width": 0.3,
                    "depth": 0.5,
                    "area": 0.15,
                    "I33": 0.003125,
                    "I22": 0.001125,
                    "J": 0.001953
                },
                "BEAM_300x600": {
                    "material": "CONCRETE",
                    "shape": "Rectangle",
                    "width": 0.3,
                    "depth": 0.6,
                    "area": 0.18,
                    "I33": 0.0054,
                    "I22": 0.00135,
                    "J": 0.002808
                },
                "BEAM_400x700": {
                    "material": "CONCRETE",
                    "shape": "Rectangle",
                    "width": 0.4,
                    "depth": 0.7,
                    "area": 0.28,
                    "I33": 0.011433,
                    "I22": 0.003733,
                    "J": 0.005947
                }
            })
        else:  # steel sections
            sections.update({
                "HEB200": {
                    "material": "STEEL",
                    "shape": "I",
                    "width": 0.2,
                    "depth": 0.2,
                    "area": 0.0078,
                    "I33": 0.00005696,
                    "I22": 0.00002003,
                    "J": 0.0000347,
                    "flange_thickness": 0.015,
                    "web_thickness": 0.009
                },
                "HEB240": {
                    "material": "STEEL",
                    "shape": "I",
                    "width": 0.24,
                    "depth": 0.24,
                    "area": 0.0106,
                    "I33": 0.00011259,
                    "I22": 0.00003923,
                    "J": 0.0000618,
                    "flange_thickness": 0.017,
                    "web_thickness": 0.01
                },
                "HEB300": {
                    "material": "STEEL",
                    "shape": "I", 
                    "width": 0.3,
                    "depth": 0.3,
                    "area": 0.0149,
                    "I33": 0.00025166,
                    "I22": 0.00008563,
                    "J": 0.000134,
                    "flange_thickness": 0.019,
                    "web_thickness": 0.011
                },
                "IPE300": {
                    "material": "STEEL",
                    "shape": "I",
                    "width": 0.15,
                    "depth": 0.3,
                    "area": 0.0054,
                    "I33": 0.00008356,
                    "I22": 0.000604,
                    "J": 0.0000204,
                    "flange_thickness": 0.0107,
                    "web_thickness": 0.0071
                },
                "IPE360": {
                    "material": "STEEL",
                    "shape": "I",
                    "width": 0.17,
                    "depth": 0.36,
                    "area": 0.0073,
                    "I33": 0.00016270,
                    "I22": 0.001043,
                    "J": 0.0000377,
                    "flange_thickness": 0.0127,
                    "web_thickness": 0.008
                },
                "IPE450": {
                    "material": "STEEL",
                    "shape": "I",
                    "width": 0.19,
                    "depth": 0.45,
                    "area": 0.0098,
                    "I33": 0.00033740,
                    "I22": 0.001676,
                    "J": 0.0000667,
                    "flange_thickness": 0.0147,
                    "web_thickness": 0.0093
                }
            })
        
        return sections

    def _get_load_patterns(self, loads: Dict) -> List[Dict]:
        """Get load patterns list"""
        return [
            {
                "name": "DEAD",
                "type": "Dead",
                "design_type": "Dead",
                "self_weight": 1.0,
                "auto_load": "None",
                "notes": f"Dead load: {loads['dead_load']} kN/m2"
            },
            {
                "name": "LIVE", 
                "type": "Live",
                "design_type": "Live",
                "self_weight": 0.0,
                "auto_load": "None",
                "notes": f"Live load: {loads['live_load']} kN/m2"
            },
            {
                "name": "WIND",
                "type": "Wind",
                "design_type": "Wind",
                "self_weight": 0.0,
                "auto_load": "None", 
                "notes": f"Wind load: {loads['wind_load']} kN/m2"
            }
        ]

async def main():
    """Main server entry point"""
    try:
        logger.info("Starting Enhanced ETABS MCP server with Excel file generation...")
        server_instance = EnhancedETABSServer()
        
        async with stdio_server() as (read_stream, write_stream):
            logger.info("Enhanced ETABS server running and waiting for connections...")
            
            try:
                await server_instance.server.run(
                    read_stream,
                    write_stream,
                    InitializationOptions(
                        server_name="enhanced-etabs-structural-design",
                        server_version="4.0.0",
                        capabilities={}
                    ),
                )
            except asyncio.CancelledError:
                logger.info("Server cancelled")
                raise
            except Exception as e:
                logger.error(f"Error during server run: {e}")
                raise
                
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Server shutdown")
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        sys.exit(1)