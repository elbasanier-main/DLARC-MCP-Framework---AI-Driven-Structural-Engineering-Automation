#!/usr/bin/env python3
"""
Simple ETABS MCP Server - Working Version
"""

import asyncio
import json
import sys
import logging
import math
import os
from typing import Any, Dict, List
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

try:
    from mcp.server import Server
    from mcp.server.models import InitializationOptions
    from mcp.server.stdio import stdio_server
    from mcp.types import Tool, TextContent
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError as e:
    logger.error(f"Import error: {e}")
    sys.exit(1)

class SimpleETABSServer:
    def __init__(self):
        self.server = Server("simple-etabs-server")
        self.building_data = {}
        self.output_dir = "etabs_exports"
        os.makedirs(self.output_dir, exist_ok=True)
        self.setup_tools()
        logger.info("Simple ETABS server initialized")

    def setup_tools(self):
        @self.server.list_tools()
        async def handle_list_tools() -> List[Tool]:
            return [
                Tool(
                    name="test_connection",
                    description="Test MCP server connection",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "message": {"type": "string", "default": "Hello from ETABS MCP!"}
                        }
                    }
                ),
                Tool(
                    name="design_building",
                    description="Design building structure",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "building_name": {"type": "string"},
                            "stories": {"type": "integer", "minimum": 1, "maximum": 50},
                            "plan_dimensions": {
                                "type": "object",
                                "properties": {
                                    "length": {"type": "number"},
                                    "width": {"type": "number"}
                                },
                                "required": ["length", "width"]
                            },
                            "bay_spacing": {
                                "type": "object",
                                "properties": {
                                    "x_direction": {"type": "number", "default": 6.0},
                                    "y_direction": {"type": "number", "default": 6.0}
                                }
                            },
                            "material": {"type": "string", "enum": ["steel", "concrete"], "default": "concrete"}
                        },
                        "required": ["building_name", "stories", "plan_dimensions"]
                    }
                ),
                Tool(
                    name="export_etabs_excel", 
                    description="Export to ETABS Excel format",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "building_name": {"type": "string"}
                        },
                        "required": ["building_name"]
                    }
                )
            ]

        @self.server.call_tool()
        async def handle_call_tool(name: str, arguments: Dict[str, Any]) -> List[TextContent]:
            logger.info(f"Tool called: {name}")
            
            if name == "test_connection":
                return [TextContent(type="text", text=f"Connection successful! Message: {arguments.get('message', 'Hello')}")]
            
            elif name == "design_building":
                building_name = arguments["building_name"]
                stories = arguments["stories"]
                length = arguments["plan_dimensions"]["length"]
                width = arguments["plan_dimensions"]["width"]
                bay_x = arguments.get("bay_spacing", {}).get("x_direction", 6.0)
                bay_y = arguments.get("bay_spacing", {}).get("y_direction", 6.0)
                material = arguments.get("material", "concrete")
                
                # Calculate structure
                x_bays = math.ceil(length / bay_x)
                y_bays = math.ceil(width / bay_y)
                total_joints = (x_bays + 1) * (y_bays + 1) * (stories + 1)
                total_columns = (x_bays + 1) * (y_bays + 1) * stories
                total_beams = stories * ((x_bays * (y_bays + 1)) + (y_bays * (x_bays + 1)))
                
                # Store building data
                self.building_data[building_name] = {
                    "name": building_name,
                    "stories": stories,
                    "dimensions": {"length": length, "width": width},
                    "bays": {"x": x_bays, "y": y_bays},
                    "material": material,
                    "summary": {
                        "joints": total_joints,
                        "columns": total_columns,
                        "beams": total_beams
                    }
                }
                
                result = f"""Building Design Complete!

Building: {building_name}
Dimensions: {length}m x {width}m x {stories * 3.5}m high
Stories: {stories}
Material: {material}
Grid Layout: {x_bays} x {y_bays} bays

Structural Elements:
- Total Joints: {total_joints}
- Columns: {total_columns}
- Beams: {total_beams}

Ready for ETABS export!"""
                
                return [TextContent(type="text", text=result)]
            
            elif name == "export_etabs_excel":
                building_name = arguments["building_name"]
                
                if building_name not in self.building_data:
                    return [TextContent(type="text", text=f"Building '{building_name}' not found. Design it first!")]
                
                building = self.building_data[building_name]
                
                # Create Excel file
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Building_Summary"
                
                # Add headers
                headers = ["Property", "Value"]
                ws.append(headers)
                
                # Add building data
                ws.append(["Building Name", building["name"]])
                ws.append(["Stories", building["stories"]])
                ws.append(["Length (m)", building["dimensions"]["length"]])
                ws.append(["Width (m)", building["dimensions"]["width"]])
                ws.append(["Material", building["material"]])
                ws.append(["Total Joints", building["summary"]["joints"]])
                ws.append(["Total Columns", building["summary"]["columns"]])
                ws.append(["Total Beams", building["summary"]["beams"]])
                
                # Style headers
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Save file
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{building_name}_ETABS_{timestamp}.xlsx"
                filepath = os.path.join(self.output_dir, filename)
                wb.save(filepath)
                
                result = f"""Excel File Created Successfully!

File: {filename}
Location: {os.path.abspath(filepath)}
Size: {os.path.getsize(filepath) / 1024:.1f} KB

Building Summary:
- {building['summary']['joints']} joints
- {building['summary']['columns']} columns  
- {building['summary']['beams']} beams

Ready for ETABS import!"""
                
                return [TextContent(type="text", text=result)]
            
            else:
                return [TextContent(type="text", text=f"Unknown tool: {name}")]

async def main():
    try:
        logger.info("Starting Simple ETABS MCP server...")
        server_instance = SimpleETABSServer()
        
        async with stdio_server() as (read_stream, write_stream):
            logger.info("Server running...")
            
            await server_instance.server.run(
                read_stream,
                write_stream,
                InitializationOptions(
                    server_name="simple-etabs-server",
                    server_version="1.0.0",
                    capabilities={}
                ),
            )
    except KeyboardInterrupt:
        logger.info("Server stopped")
    except Exception as e:
        logger.error(f"Server error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    asyncio.run(main())