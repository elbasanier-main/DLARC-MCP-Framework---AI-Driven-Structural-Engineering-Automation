#!/usr/bin/env python3
"""
Fixed ETABS Excel MCP Server - Corrected Syntax (No Emojis in Code)
Ready for ETABS 21 import with proper database table format
"""

import asyncio
import json
import sys
import logging
import math
import os
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

try:
    from mcp.server import Server
    from mcp.server.models import InitializationOptions
    from mcp.server.stdio import stdio_server
    from mcp.types import Tool, TextContent
    
    # Excel generation libraries
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
except ImportError as e:
    logger.error(f"Import error: {e}")
    logger.error("Make sure you have required packages: pip install mcp openpyxl")
    sys.exit(1)

class FixedETABSServer:
    def __init__(self):
        try:
            self.server = Server("fixed-etabs-structural-design")
            self.building_data = {}
            self.output_dir = "etabs_exports"
            self.setup_output_directory()
            self.setup_tools()
            logger.info("Server initialized successfully - Fixed ETABS MCP Server")
        except Exception as e:
            logger.error(f"Server initialization failed: {e}")
            raise

    def setup_output_directory(self):
        """Create output directory for exported files"""
        try:
            os.makedirs(self.output_dir, exist_ok=True)
            logger.info(f"Output directory ready: {self.output_dir}")
        except Exception as e:
            logger.error(f"Failed to create output directory: {e}")

    def setup_tools(self):
        """Setup ETABS tools with corrected format"""
        
        @self.server.list_tools()
        async def handle_list_tools() -> List[Tool]:
            return [
                Tool(
                    name="test_connection",
                    description="Test MCP server connection",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "message": {"type": "string", "description": "Test message", "default": "Hello from Fixed ETABS MCP!"}
                        }
                    }
                ),
                
                Tool(
                    name="design_building",
                    description="Design building with correct ETABS format",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "building_name": {"type": "string", "description": "Building name/identifier"},
                            "stories": {"type": "integer", "description": "Number of stories", "minimum": 1, "maximum": 50},
                            "story_height": {"type": "number", "description": "Typical story height (m)", "default": 3.5},
                            "plan_dimensions": {
                                "type": "object",
                                "properties": {
                                    "length": {"type": "number", "description": "Building length (m)"},
                                    "width": {"type": "number", "description": "Building width (m)"}
                                },
                                "required": ["length", "width"]
                            },
                            "bay_spacing": {
                                "type": "object",
                                "properties": {
                                    "x_direction": {"type": "number", "description": "Bay spacing in X direction (m)", "default": 6.0},
                                    "y_direction": {"type": "number", "description": "Bay spacing in Y direction (m)", "default": 6.0}
                                }
                            },
                            "material": {"type": "string", "enum": ["steel", "concrete"], "default": "concrete"}
                        },
                        "required": ["building_name", "stories", "plan_dimensions"]
                    }
                ),
                
                Tool(
                    name="export_etabs_excel",
                    description="Export to ETABS 21 compatible Excel format",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "building_name": {"type": "string", "description": "Building name to export"}
                        },
                        "required": ["building_name"]
                    }
                )
            ]

        @self.server.call_tool()
        async def handle_call_tool(name: str, arguments: Dict[str, Any]) -> List[TextContent]:
            try:
                logger.info(f"Tool called: {name} with arguments: {arguments}")
                
                if name == "test_connection":
                    return self._test_connection(arguments)
                elif name == "design_building":
                    return self._design_building(arguments)
                elif name == "export_etabs_excel":
                    return self._export_etabs_excel(arguments)
                else:
                    return [TextContent(type="text", text=f"Unknown tool: {name}")]
                    
            except Exception as e:
                error_msg = f"Error in tool '{name}': {str(e)}"
                logger.error(error_msg)
                return [TextContent(type="text", text=error_msg)]

    def _test_connection(self, arguments: Dict[str, Any]) -> List[TextContent]:
        """Test connection tool"""
        message = arguments.get("message", "Hello from Fixed ETABS MCP!")
        response = f"""ETABS MCP Server connected successfully!

Message: {message}

Server Status: Running
Available Tools: 3 (test_connection, design_building, export_etabs_excel)

FIXES APPLIED:
- Fixed syntax errors (removed unicode characters)
- Ready for direct import to ETABS 21
- Proper database table format
- Correct column headers

Output Directory: {os.path.abspath(self.output_dir)}"""
        
        return [TextContent(type="text", text=response)]

    def _design_building(self, arguments: Dict[str, Any]) -> List[TextContent]:
        """Design building with proper ETABS structure"""
        building_name = arguments["building_name"]
        stories = arguments["stories"]
        story_height = arguments.get("story_height", 3.5)
        plan_dims = arguments["plan_dimensions"]
        bay_spacing = arguments.get("bay_spacing", {"x_direction": 6.0, "y_direction": 6.0})
        material = arguments.get("material", "concrete")
        
        # Calculate building geometry
        length = plan_dims["length"]
        width = plan_dims["width"]
        total_height = stories * story_height
        
        # Calculate grid lines
        x_bays = math.ceil(length / bay_spacing["x_direction"])
        y_bays = math.ceil(width / bay_spacing["y_direction"])
        
        actual_x_spacing = length / x_bays
        actual_y_spacing = width / y_bays
        
        # Generate joints with ETABS naming convention
        joints = []
        joint_id = 1
        joint_map = {}  # To track joints by (story, i, j)
        
        for story in range(stories + 1):  # Include base level
            z = story * story_height
            for i in range(x_bays + 1):
                for j in range(y_bays + 1):
                    # Create unique joint ID
                    joint_name = str(joint_id)
                    joint_map[(story, i, j)] = joint_name
                    
                    joints.append({
                        "Joint": joint_name,
                        "CoordSys": "GLOBAL",
                        "CoordType": "Cartesian", 
                        "XorR": round(i * actual_x_spacing, 3),
                        "Y": round(j * actual_y_spacing, 3),
                        "Z": round(z, 3),
                        "SpecialJt": "No",
                        "GlobalX": round(i * actual_x_spacing, 3),
                        "GlobalY": round(j * actual_y_spacing, 3),
                        "GlobalZ": round(z, 3),
                        "story": story,
                        "grid_i": i,
                        "grid_j": j
                    })
                    joint_id += 1
        
        # Generate frames with ETABS naming
        frames = []
        frame_id = 1
        
        # Columns
        for story in range(stories):
            for i in range(x_bays + 1):
                for j in range(y_bays + 1):
                    # Find joints for this column
                    joint_bottom = joint_map[(story, i, j)]
                    joint_top = joint_map[(story + 1, i, j)]
                    
                    # ETABS style frame naming
                    frame_name = f"C{frame_id}"
                    
                    frames.append({
                        "Frame": frame_name,
                        "JointI": joint_bottom,
                        "JointJ": joint_top,  
                        "Length": story_height,
                        "MatProp": "CONC" if material == "concrete" else "STEEL",
                        "Section": "COL1",
                        "Angle": 0,
                        "CardinalPt": "10-Centroid",
                        "type": "column"
                    })
                    frame_id += 1
        
        # Beams - X direction
        for story in range(1, stories + 1):
            for j in range(y_bays + 1):
                for i in range(x_bays):
                    joint_i = joint_map[(story, i, j)]
                    joint_j = joint_map[(story, i + 1, j)]
                    frame_name = f"B{frame_id}"
                    
                    frames.append({
                        "Frame": frame_name,
                        "JointI": joint_i,
                        "JointJ": joint_j,
                        "Length": actual_x_spacing,
                        "MatProp": "CONC" if material == "concrete" else "STEEL",
                        "Section": "BEAM1",
                        "Angle": 0,
                        "CardinalPt": "10-Centroid",
                        "type": "beam"
                    })
                    frame_id += 1
        
        # Beams - Y direction  
        for story in range(1, stories + 1):
            for i in range(x_bays + 1):
                for j in range(y_bays):
                    joint_i = joint_map[(story, i, j)]
                    joint_j = joint_map[(story, i, j + 1)]
                    frame_name = f"B{frame_id}"
                    
                    frames.append({
                        "Frame": frame_name,
                        "JointI": joint_i,
                        "JointJ": joint_j,
                        "Length": actual_y_spacing,
                        "MatProp": "CONC" if material == "concrete" else "STEEL",
                        "Section": "BEAM1",
                        "Angle": 0,
                        "CardinalPt": "10-Centroid",
                        "type": "beam"
                    })
                    frame_id += 1
        
        # Generate materials in ETABS format
        materials = self._get_etabs_materials(material)
        sections = self._get_etabs_sections(material)
        
        # Store building data
        building_data = {
            "name": building_name,
            "geometry": {
                "stories": stories,
                "story_height": story_height,
                "total_height": total_height,
                "plan_dimensions": {"length": length, "width": width},
                "bay_spacing": {"x": actual_x_spacing, "y": actual_y_spacing}
            },
            "material": material,
            "joints": joints,
            "frames": frames,
            "materials": materials,
            "sections": sections,
            "design_summary": {
                "total_joints": len(joints),
                "total_frames": len(frames),
                "columns": len([f for f in frames if f["type"] == "column"]),
                "beams": len([f for f in frames if f["type"] == "beam"])
            }
        }
        
        self.building_data[building_name] = building_data
        
        result = {
            "building_summary": {
                "name": building_name,
                "dimensions": f"{length}m x {width}m x {total_height}m",
                "stories": stories,
                "material": material
            },
            "model_statistics": building_data["design_summary"],
            "etabs_format": "Fixed and ready for ETABS 21 import",
            "next_steps": [
                f"Use 'export_etabs_excel' to generate Excel file",
                f"File will use correct ETABS table names and column headers"
            ]
        }
        
        response = f"Building Design Complete - ETABS Format Fixed:\n\n{json.dumps(result, indent=2)}"
        return [TextContent(type="text", text=response)]

    def _export_etabs_excel(self, arguments: Dict[str, Any]) -> List[TextContent]:
        """Export with corrected ETABS format"""
        building_name = arguments["building_name"]
        
        if building_name not in self.building_data:
            return [TextContent(type="text", text=f"Building '{building_name}' not found. Please design it first.")]
        
        building = self.building_data[building_name]
        
        # Create Excel file with correct ETABS format
        file_path = self._create_etabs_excel_file(building)
        
        response = f"""ETABS Excel File Generated Successfully!

File Details:
  - File Name: {os.path.basename(file_path)}
  - Full Path: {os.path.abspath(file_path)}
  - File Size: {os.path.getsize(file_path) / 1024:.1f} KB
  - Format: ETABS 21 database table structure

Model Statistics:
  - Building: {building['name']}
  - Joints: {building['design_summary']['total_joints']}
  - Frames: {building['design_summary']['total_frames']}
  - Columns: {building['design_summary']['columns']}
  - Beams: {building['design_summary']['beams']}

Fixed Features:
  - Proper table names (e.g., "Joint Coordinates", "Connectivity - Frame")
  - Correct column headers matching ETABS exactly
  - Proper data types and formatting
  - Compatible with ETABS 21 import process

ETABS 21 Import Steps:
1. Open ETABS 21
2. File > Import > Excel File
3. Select: {os.path.basename(file_path)}
4. Map tables automatically (names should match)
5. Import and verify

Excel Worksheets:
  - Joint Coordinates - Joint locations and properties
  - Connectivity - Frame - Frame connectivity and properties
  - Material Properties 02 - Basic Mechanical Properties
  - Frame Section Properties 01 - General

Ready for direct import!"""
        
        return [TextContent(type="text", text=response)]

    def _create_etabs_excel_file(self, building: Dict) -> str:
        """Create Excel file with correct ETABS format"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{building['name']}_ETABS_Fixed_{timestamp}.xlsx"
        file_path = os.path.join(self.output_dir, filename)
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create worksheets with EXACT ETABS table names
        self._create_joint_coordinates_sheet(wb, building)
        self._create_connectivity_frame_sheet(wb, building)
        self._create_material_properties_sheet(wb, building)
        self._create_frame_section_properties_sheet(wb, building)
        
        # Save file
        wb.save(file_path)
        logger.info(f"Excel file created: {file_path}")
        return file_path

    def _create_joint_coordinates_sheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create Joint Coordinates sheet with exact ETABS format"""
        ws = wb.create_sheet("Joint Coordinates")
        
        # EXACT ETABS column headers
        headers = [
            "Joint", "CoordSys", "CoordType", "XorR", "Y", "Z", 
            "SpecialJt", "GlobalX", "GlobalY", "GlobalZ"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add joint data in ETABS format
        for joint in building["joints"]:
            row = [
                joint["Joint"],           # Joint name
                joint["CoordSys"],        # "GLOBAL"
                joint["CoordType"],       # "Cartesian"
                joint["XorR"],            # X coordinate
                joint["Y"],               # Y coordinate  
                joint["Z"],               # Z coordinate
                joint["SpecialJt"],       # "No"
                joint["GlobalX"],         # Global X (same as XorR)
                joint["GlobalY"],         # Global Y (same as Y)
                joint["GlobalZ"]          # Global Z (same as Z)
            ]
            ws.append(row)
        
        # Auto-size columns
        self._autosize_columns(ws)

    def _create_connectivity_frame_sheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create Connectivity - Frame sheet with exact ETABS format"""
        ws = wb.create_sheet("Connectivity - Frame")
        
        # EXACT ETABS column headers
        headers = [
            "Frame", "JointI", "JointJ", "Length", "MatProp", "Section", 
            "Angle", "CardinalPt", "NonPrismatic", "DesignOrient", "GUID"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add frame data
        for frame in building["frames"]:
            row = [
                frame["Frame"],           # Frame name
                frame["JointI"],          # Start joint
                frame["JointJ"],          # End joint
                frame["Length"],          # Length
                frame["MatProp"],         # Material property
                frame["Section"],         # Section property
                frame["Angle"],           # Angle (0)
                frame.get("CardinalPt", "10-Centroid"),  # Cardinal point
                "No",                     # NonPrismatic
                "Global",                 # DesignOrient
                ""                        # GUID (empty)
            ]
            ws.append(row)
        
        # Auto-size columns
        self._autosize_columns(ws)

    def _create_material_properties_sheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create Material Properties 02 - Basic Mechanical Properties sheet"""
        ws = wb.create_sheet("Material Properties 02 - Basic Mechanical Properties")
        
        # EXACT ETABS column headers for material properties
        headers = [
            "Material", "Type", "SymType", "TempDepend", "Color", "Notes", "GUID",
            "E", "U", "A", "G"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add material data
        for mat_name, mat_props in building["materials"].items():
            row = [
                mat_name,                 # Material name
                mat_props["Type"],        # Material type
                "Isotropic",              # Symmetry type
                "No",                     # Temperature dependent
                mat_props.get("Color", "Blue"),  # Color
                "",                       # Notes
                "",                       # GUID
                mat_props["E"],           # Elastic modulus
                mat_props["U"],           # Poisson's ratio
                mat_props["A"],           # Thermal coefficient
                mat_props["G"]            # Shear modulus
            ]
            ws.append(row)
        
        # Auto-size columns
        self._autosize_columns(ws)

    def _create_frame_section_properties_sheet(self, wb: openpyxl.Workbook, building: Dict):
        """Create Frame Section Properties 01 - General sheet"""
        ws = wb.create_sheet("Frame Section Properties 01 - General")
        
        # EXACT ETABS column headers for frame sections
        headers = [
            "SectionName", "Material", "Shape", "t3", "t2", "tf", "tw", 
            "Area", "TorsConst", "I33", "I22", "I12", "AS2", "AS3", 
            "S33", "S22", "Z33", "Z22", "R33", "R22"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add section data
        for sect_name, sect_props in building["sections"].items():
            row = [
                sect_name,                        # Section name
                sect_props["Material"],           # Material
                sect_props["Shape"],              # Shape
                sect_props["t3"],                 # Depth (t3)
                sect_props["t2"],                 # Width (t2)
                sect_props.get("tf", 0),          # Flange thickness
                sect_props.get("tw", 0),          # Web thickness
                sect_props["Area"],               # Area
                sect_props["TorsConst"],          # Torsion constant
                sect_props["I33"],                # Moment of inertia 33
                sect_props["I22"],                # Moment of inertia 22
                0,                                # I12 (product of inertia)
                sect_props.get("AS2", 0),         # Shear area 2
                sect_props.get("AS3", 0),         # Shear area 3
                sect_props.get("S33", 0),         # Section modulus 33
                sect_props.get("S22", 0),         # Section modulus 22
                sect_props.get("Z33", 0),         # Plastic section modulus 33
                sect_props.get("Z22", 0),         # Plastic section modulus 22
                sect_props.get("R33", 0),         # Radius of gyration 33
                sect_props.get("R22", 0)          # Radius of gyration 22
            ]
            ws.append(row)
        
        # Auto-size columns
        self._autosize_columns(ws)

    def _autosize_columns(self, ws):
        """Auto-size columns in worksheet"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _get_etabs_materials(self, material: str) -> Dict:
        """Get materials in ETABS format"""
        if material == "concrete":
            return {
                "CONC": {
                    "Type": "Concrete",
                    "E": 30000000,      # kN/m² (30 GPa)
                    "U": 0.2,           # Poisson's ratio
                    "A": 1e-5,          # Thermal coefficient
                    "G": 12500000,      # Shear modulus
                    "Color": "Gray"
                }
            }
        else:  # steel
            return {
                "STEEL": {
                    "Type": "Steel",
                    "E": 200000000,     # kN/m² (200 GPa)
                    "U": 0.3,           # Poisson's ratio
                    "A": 1.2e-5,        # Thermal coefficient
                    "G": 76923000,      # Shear modulus
                    "Color": "Blue"
                }
            }

    def _get_etabs_sections(self, material: str) -> Dict:
        """Get sections in ETABS format"""
        if material == "concrete":
            return {
                "COL1": {
                    "Material": "CONC",
                    "Shape": "Rectangular",
                    "t3": 0.4,          # Depth (400mm)
                    "t2": 0.4,          # Width (400mm)
                    "Area": 0.16,       # m²
                    "TorsConst": 0.003413,  # m⁴
                    "I33": 0.002133,    # m⁴
                    "I22": 0.002133,    # m⁴
                    "AS2": 0.133,       # Shear area
                    "AS3": 0.133,       # Shear area
                    "S33": 0.010667,    # Section modulus
                    "S22": 0.010667,    # Section modulus
                    "Z33": 0.016,       # Plastic section modulus
                    "Z22": 0.016,       # Plastic section modulus
                    "R33": 0.115,       # Radius of gyration
                    "R22": 0.115        # Radius of gyration
                },
                "BEAM1": {
                    "Material": "CONC",
                    "Shape": "Rectangular",
                    "t3": 0.6,          # Depth (600mm)
                    "t2": 0.3,          # Width (300mm)
                    "Area": 0.18,       # m²
                    "TorsConst": 0.002808,  # m⁴
                    "I33": 0.0054,      # m⁴
                    "I22": 0.00135,     # m⁴
                    "AS2": 0.15,        # Shear area
                    "AS3": 0.15,        # Shear area
                    "S33": 0.018,       # Section modulus
                    "S22": 0.009,       # Section modulus
                    "Z33": 0.027,       # Plastic section modulus
                    "Z22": 0.0135,      # Plastic section modulus
                    "R33": 0.173,       # Radius of gyration
                    "R22": 0.087        # Radius of gyration
                }
            }
        else:  # steel
            return {
                "COL1": {
                    "Material": "STEEL",
                    "Shape": "I/Wide Flange",
                    "t3": 0.24,         # Depth (HEB240)
                    "t2": 0.24,         # Width
                    "tf": 0.017,        # Flange thickness
                    "tw": 0.01,         # Web thickness
                    "Area": 0.0106,     # m²
                    "TorsConst": 0.0000618,  # m⁴
                    "I33": 0.00011259,  # m⁴
                    "I22": 0.00003923,  # m⁴
                    "AS2": 0.0024,      # Shear area
                    "AS3": 0.0024,      # Shear area
                    "S33": 0.000938,    # Section modulus
                    "S22": 0.000327,    # Section modulus
                    "Z33": 0.001056,    # Plastic section modulus
                    "Z22": 0.000491,    # Plastic section modulus
                    "R33": 0.103,       # Radius of gyration
                    "R22": 0.061        # Radius of gyration
                },
                "BEAM1": {
                    "Material": "STEEL",
                    "Shape": "I/Wide Flange",
                    "t3": 0.36,         # Depth (IPE360)
                    "t2": 0.17,         # Width
                    "tf": 0.0127,       # Flange thickness
                    "tw": 0.008,        # Web thickness
                    "Area": 0.0073,     # m²
                    "TorsConst": 0.0000377,  # m⁴
                    "I33": 0.0001627,   # m⁴
                    "I22": 0.000001043, # m⁴
                    "AS2": 0.002016,    # Shear area
                    "AS3": 0.002016,    # Shear area
                    "S33": 0.000904,    # Section modulus
                    "S22": 0.000123,    # Section modulus
                    "Z33": 0.001019,    # Plastic section modulus
                    "Z22": 0.000189,    # Plastic section modulus
                    "R33": 0.149,       # Radius of gyration
                    "R22": 0.038        # Radius of gyration
                }
            }

async def main():
    """Main server entry point"""
    try:
        logger.info("Starting Fixed ETABS MCP server...")
        server_instance = FixedETABSServer()
        
        async with stdio_server() as (read_stream, write_stream):
            logger.info("Fixed ETABS server running and waiting for connections...")
            
            try:
                await server_instance.server.run(
                    read_stream,
                    write_stream,
                    InitializationOptions(
                        server_name="fixed-etabs-structural-design",
                        server_version="6.0.0",
                        capabilities={}
                    ),
                )
            except asyncio.CancelledError:
                logger.info("Server cancelled")
                raise
            except Exception as e:
                logger.error(f"Error during server run: {e}")
                raise
                
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Server shutdown")
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        sys.exit(1)